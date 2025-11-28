import openpyxl
from datetime import datetime

excel_path = r"C:\Users\Duncan\OneDrive\Documents\Yewtree Square Stuff\OMC\Invoices\2025 Invoices.xlsx"

print("\n" + "="*80)
print("ANALYZING: 2025 Invoices.xlsx")
print("="*80)

try:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    print(f"\nWorkbook contains {len(wb.sheetnames)} sheet(s):")
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")
    
    # Analyze each sheet
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print("\n" + "="*80)
        print(f"SHEET: {sheet_name}")
        print("="*80)
        
        # Get dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column
        print(f"\nDimensions: {max_row} rows x {max_col} columns")
        
        # Show first 10 rows to understand structure
        print("\nFirst 10 rows:")
        print("-"*80)
        
        for row_idx in range(1, min(11, max_row + 1)):
            row_values = []
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                if value is not None:
                    if isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d')
                    elif isinstance(value, (int, float)):
                        value = f"{value:,.2f}" if isinstance(value, float) else str(value)
                    row_values.append(str(value)[:30])
                else:
                    row_values.append("")
            
            print(f"Row {row_idx:2}: {' | '.join(row_values)}")
        
        if max_row > 10:
            print(f"\n... and {max_row - 10} more rows")
        
        # Try to identify header row and data structure
        print("\n" + "-"*80)
        print("Column Headers (Row 1):")
        for col_idx in range(1, max_col + 1):
            header = sheet.cell(row=1, column=col_idx).value
            if header:
                print(f"  Col {col_idx}: {header}")
    
    wb.close()
    
except Exception as e:
    print(f"\nERROR: {str(e)}")
    import traceback
    traceback.print_exc()
