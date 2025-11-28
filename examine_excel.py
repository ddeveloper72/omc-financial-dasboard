"""
Examine Excel file structure to understand the data format
"""
import openpyxl
from pathlib import Path

# Path to the Excel file
excel_path = r"C:\Users\Duncan\OneDrive\Documents\Yewtree Square Stuff\OMC\Costs\YES TREE SQUARE COSTINGS 2025.xlsx"

# Load the workbook
wb = openpyxl.load_workbook(excel_path, data_only=True)

print(f"Excel file: {excel_path}")
print(f"\nSheet names: {wb.sheetnames}")
print("=" * 80)

# Examine each sheet
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n\nSheet: {sheet_name}")
    print("-" * 80)
    
    # Get dimensions
    print(f"Max row: {sheet.max_row}, Max column: {sheet.max_column}")
    
    # Print first 20 rows to understand structure
    print("\nFirst 20 rows:")
    for row_idx in range(1, min(21, sheet.max_row + 1)):
        row_values = []
        for col_idx in range(1, min(15, sheet.max_column + 1)):  # First 15 columns
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is not None:
                row_values.append(f"Col{col_idx}: {value}")
        
        if row_values:
            print(f"Row {row_idx}: {' | '.join(row_values)}")

wb.close()
