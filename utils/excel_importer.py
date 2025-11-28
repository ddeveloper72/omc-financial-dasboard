"""
Excel Importer - YTS Budget Analysis

This module imports actual cost data from Excel workbooks containing
supplier invoices and receipts.

EU AI Act Compliance:
- Rule-based deterministic parsing (no AI/ML algorithms)
- Transparent category mapping logic
- Audit trail maintained via database logging

Data Processing:
- Reads Excel files using openpyxl (https://openpyxl.readthedocs.io/)
- Maps supplier sheets to budget categories
- Handles duplicate prevention via database constraints
- Logs import results for transparency

For full compliance documentation, see COMPLIANCE.md
"""

import openpyxl
from datetime import datetime
from models import db, ActualCost, ChargeCategory, Document
from sqlalchemy.exc import IntegrityError

# Supplier to category mapping
# Maps Excel sheet names to ChargeCategory IDs based on service type
SUPPLIER_CATEGORY_MAP = {
    'GATES ': 8,  # Security (gate maintenance)
    'BIN ROOMS ': 4,  # Cleaning (waste management infrastructure)
    'FIRE ALARM CALL OUTS ': 8,  # Security (fire safety)
    'ASECENION LIFTS ': 6,  # Lift/Elevator
    'WILSHIRE MAINTANCE CALL OUTS': 3,  # Maintenance
    'HAWKEYE': 8,  # Security (pest control)
    'BOARD GAIS ': 5,  # Utilities (gas/heating)
    'GARDEN ': 7,  # Grounds
    'OMEGA ': 8,  # Security (fire safety survey)
    'D&G CLEANING ': 4,  # Cleaning
    'MAGNET ': 5,  # Utilities (electricity)
    'THORNTONS ': 4,  # Cleaning (refuse collection)
}

def parse_excel_date(cell_value):
    """
    Parse date from Excel cell (handles datetime objects and DD/MM/YYYY strings).
    
    Args:
        cell_value: Cell value which may be datetime object or string
        
    Returns:
        datetime.date object or None if unparseable
    """
    if cell_value is None:
        return None
    
    # If already a datetime object
    if isinstance(cell_value, datetime):
        return cell_value.date()
    
    # If it's a string, try parsing DD/MM/YYYY format
    if isinstance(cell_value, str):
        try:
            # Try DD/MM/YYYY format
            dt = datetime.strptime(cell_value, '%d/%m/%Y')
            return dt.date()
        except ValueError:
            try:
                # Try alternative format YYYY-MM-DD
                dt = datetime.strptime(cell_value, '%Y-%m-%d')
                return dt.date()
            except ValueError:
                return None
    
    return None


def is_header_or_total_row(row_cells):
    """
    Check if row is a header or totals row (should be skipped).
    
    Args:
        row_cells: List of cell values from a row
        
    Returns:
        True if row should be skipped, False otherwise
    """
    if not row_cells or all(cell is None for cell in row_cells):
        return True
    
    first_cell = str(row_cells[0]).strip().upper() if row_cells[0] else ''
    
    # Skip headers and totals
    skip_keywords = ['SUPPLIER', 'DATE', 'INVOICE', 'TOTAL', 'TOTALS', 'SUGGESTION', 'NEW COMPANY', 'REFUND']
    
    for keyword in skip_keywords:
        if keyword in first_cell:
            return True
    
    return False


def extract_invoice_data(row_cells, sheet_name):
    """
    Extract invoice data from Excel row.
    
    Args:
        row_cells: List of cell values from a row
        sheet_name: Name of the Excel sheet (supplier name)
        
    Returns:
        Dictionary with invoice data or None if invalid row
    """
    # Skip empty rows or header/total rows
    if is_header_or_total_row(row_cells):
        return None
    
    # Expected columns: SUPPLIER, DATE, INVOICE, NET, VAT, TOTAL, [DESCRIPTION]
    # Some rows have missing columns (Col3/invoice missing)
    
    # Extract supplier (Col1)
    supplier = str(row_cells[0]).strip() if row_cells[0] else sheet_name.strip()
    if not supplier or supplier.upper() in ['SUPPLIER', 'NONE']:
        return None
    
    # Extract date (Col2)
    invoice_date = parse_excel_date(row_cells[1]) if len(row_cells) > 1 else None
    if not invoice_date:
        return None
    
    # Extract invoice number (Col3) - may be missing
    invoice_number = str(row_cells[2]).strip() if len(row_cells) > 2 and row_cells[2] else ''
    
    # Extract amounts (Col4=NET, Col5=VAT, Col6=TOTAL)
    try:
        net_amount = float(row_cells[3]) if len(row_cells) > 3 and row_cells[3] else 0.0
        vat_amount = float(row_cells[4]) if len(row_cells) > 4 and row_cells[4] else 0.0
        total_amount = float(row_cells[5]) if len(row_cells) > 5 and row_cells[5] else 0.0
    except (ValueError, TypeError):
        return None
    
    # Skip rows with zero total
    if total_amount == 0.0:
        return None
    
    # Extract description (Col7)
    description = str(row_cells[6]).strip() if len(row_cells) > 6 and row_cells[6] else ''
    
    return {
        'supplier': supplier,
        'date': invoice_date,
        'invoice_number': invoice_number,
        'description': description,
        'net_amount': net_amount,
        'vat_amount': vat_amount,
        'total_amount': total_amount,
        'year': invoice_date.year,
        'month': invoice_date.month
    }


def import_excel_file(excel_path, document_id=None):
    """
    Import actual costs from Excel file.
    
    Args:
        excel_path: Path to Excel workbook
        document_id: Optional Document ID to link costs to source document
        
    Returns:
        Dictionary with import statistics
    """
    stats = {
        'total_rows': 0,
        'imported': 0,
        'skipped': 0,
        'duplicates': 0,
        'errors': 0,
        'sheets_processed': 0
    }
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            stats['sheets_processed'] += 1
            
            # Get category ID for this supplier/sheet
            category_id = SUPPLIER_CATEGORY_MAP.get(sheet_name)
            if not category_id:
                print(f"Warning: No category mapping for sheet '{sheet_name}', skipping")
                continue
            
            print(f"\nProcessing sheet: {sheet_name} (Category ID: {category_id})")
            
            # Iterate through rows (skip first row which is often header)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                stats['total_rows'] += 1
                
                # Extract invoice data
                invoice_data = extract_invoice_data(row, sheet_name)
                if not invoice_data:
                    stats['skipped'] += 1
                    continue
                
                # Check for existing record (duplicate detection)
                # If invoice number exists, check by supplier + invoice + date
                # Otherwise check by supplier + date + total_amount
                if invoice_data['invoice_number']:
                    existing = ActualCost.query.filter_by(
                        supplier=invoice_data['supplier'],
                        invoice_number=invoice_data['invoice_number'],
                        date=invoice_data['date']
                    ).first()
                else:
                    existing = ActualCost.query.filter_by(
                        supplier=invoice_data['supplier'],
                        date=invoice_data['date'],
                        total_amount=invoice_data['total_amount']
                    ).first()
                
                if existing:
                    stats['duplicates'] += 1
                    continue
                
                # Create ActualCost record
                try:
                    actual_cost = ActualCost(
                        date=invoice_data['date'],
                        supplier=invoice_data['supplier'],
                        invoice_number=invoice_data['invoice_number'],
                        description=invoice_data['description'],
                        net_amount=invoice_data['net_amount'],
                        vat_amount=invoice_data['vat_amount'],
                        total_amount=invoice_data['total_amount'],
                        currency='EUR',
                        year=invoice_data['year'],
                        month=invoice_data['month'],
                        category_id=category_id,
                        document_id=document_id,
                        source_sheet=sheet_name
                    )
                    
                    db.session.add(actual_cost)
                    stats['imported'] += 1
                    
                except Exception as e:
                    db.session.rollback()
                    stats['errors'] += 1
                    print(f"Error on row {row_idx}: {e}")
            
        # Commit all changes
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        print(f"Fatal error: {e}")
        stats['errors'] += 1
    
    return stats


def print_import_summary(stats):
    """Print formatted import summary"""
    print("\n" + "="*60)
    print("IMPORT SUMMARY")
    print("="*60)
    print(f"Sheets processed: {stats['sheets_processed']}")
    print(f"Total rows examined: {stats['total_rows']}")
    print(f"Successfully imported: {stats['imported']}")
    print(f"Skipped (invalid/empty): {stats['skipped']}")
    print(f"Duplicates (already imported): {stats['duplicates']}")
    print(f"Errors: {stats['errors']}")
    print("="*60)
